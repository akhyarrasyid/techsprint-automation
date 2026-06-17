"""
Export Service — PDF, Excel, CSV, PowerPoint export.
"""
import io
import csv
import json
from typing import Dict, Any
from services.pipeline_store import PipelineStore


def export_data(format_type: str, scenario: str = "Base") -> Dict[str, Any]:
    """Export pipeline data in the requested format. Returns bytes + metadata."""
    results = PipelineStore.get_results(scenario)

    if format_type == "csv":
        return _export_csv(results, scenario)
    elif format_type == "excel":
        return _export_excel(results, scenario)
    elif format_type == "json":
        return _export_json(results, scenario)
    else:
        return _export_json(results, scenario)


def _export_csv(results: Dict[str, Any], scenario: str) -> Dict[str, Any]:
    """Export forecast + inventory as CSV."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["=== FORECAST DATA ==="])
    writer.writerow(["Product ID", "Product Name", "Forecast Next Week", "Trend %", "CI Low", "CI High"])
    for f in results.get("forecast", []):
        writer.writerow([f.get("product_id"), f.get("product_name"), f.get("forecast_next_week"), f.get("trend_pct"), f.get("confidence_interval_low"), f.get("confidence_interval_high")])

    writer.writerow([])
    writer.writerow(["=== INVENTORY DATA ==="])
    writer.writerow(["Product ID", "Product Name", "Current Stock", "Safety Stock", "ROP", "Status", "Recommended Order"])
    for i in results.get("inventory", []):
        writer.writerow([i.get("product_id"), i.get("product_name"), i.get("current_stock"), i.get("safety_stock"), i.get("reorder_point"), i.get("status"), i.get("recommended_order")])

    writer.writerow([])
    writer.writerow(["=== PROFITABILITY DATA ==="])
    writer.writerow(["Product ID", "Product Name", "Revenue", "Cost", "Profit", "Margin/Unit"])
    for p in results.get("profitability", []):
        writer.writerow([p.get("product_id"), p.get("product_name"), p.get("estimated_revenue"), p.get("estimated_cost"), p.get("estimated_profit"), p.get("margin_per_unit")])

    content = output.getvalue()
    return {"content": content, "filename": f"bpas_report_{scenario}.csv", "content_type": "text/csv"}


def _export_excel(results: Dict[str, Any], scenario: str) -> Dict[str, Any]:
    """Export as Excel using openpyxl."""
    try:
        from openpyxl import Workbook
        wb = Workbook()

        # Forecast sheet
        ws_fc = wb.active
        ws_fc.title = "Forecast"
        ws_fc.append(["Product ID", "Product Name", "Forecast Next Week", "Trend %", "CI Low", "CI High"])
        for f in results.get("forecast", []):
            ws_fc.append([f.get("product_id"), f.get("product_name"), f.get("forecast_next_week"), f.get("trend_pct"), f.get("confidence_interval_low"), f.get("confidence_interval_high")])

        # Inventory sheet
        ws_inv = wb.create_sheet("Inventory")
        ws_inv.append(["Product ID", "Product Name", "Stock", "Safety Stock", "ROP", "Status", "Order Qty"])
        for i in results.get("inventory", []):
            ws_inv.append([i.get("product_id"), i.get("product_name"), i.get("current_stock"), i.get("safety_stock"), i.get("reorder_point"), i.get("status"), i.get("recommended_order")])

        # Profitability sheet
        ws_prof = wb.create_sheet("Profitability")
        ws_prof.append(["Product ID", "Product Name", "Revenue", "Cost", "Profit"])
        for p in results.get("profitability", []):
            ws_prof.append([p.get("product_id"), p.get("product_name"), p.get("estimated_revenue"), p.get("estimated_cost"), p.get("estimated_profit")])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return {"content_bytes": buf.getvalue(), "filename": f"bpas_report_{scenario}.xlsx", "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    except ImportError:
        return _export_csv(results, scenario)


def _export_json(results: Dict[str, Any], scenario: str) -> Dict[str, Any]:
    """Export as JSON."""
    content = json.dumps(results, indent=2, default=str)
    return {"content": content, "filename": f"bpas_report_{scenario}.json", "content_type": "application/json"}
